import os
import requests
import openpyxl
from openai import OpenAI, APIError
from openpyxl.drawing.image import Image

def download_image(url, filename):
    """Download an image from URL and save it locally."""
    response = requests.get(url)
    response.raise_for_status()  # Check for errors in the response
    with open(filename, 'wb') as f:
        f.write(response.content)
    print(f"Image saved as {filename}")

def insert_image_to_excel(sheet, input_image, cell):
    """Insert an image into a specified cell in an Excel sheet."""
    img = Image(input_image)
    sheet.add_image(img, cell)

def generate_and_insert_images():
    client = OpenAI()
    client.api_key = os.getenv('OPENAI_API_KEY')
    max_attempts = 3  # Maximum number of attempts per prompt

    # Load the workbook
    wb = openpyxl.load_workbook('input.xlsx')
    prompt_cells = [f'A{i}' for i in range(2, 260)] # Change values accordingly
    image_cells = [f'B{i}' for i in range(2, 260)]
    sheets = ['Sheet Name 1', 'Sheet Name 2'] # Change names accordingly

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]

        for prompt_cell, image_cell in zip(prompt_cells, image_cells):
            prompt = sheet[prompt_cell].value
            if prompt:  # Ensure there is a prompt
                attempt = 0
                while attempt < max_attempts:
                    try:
                        print(f"Attempt {attempt + 1}: Processing prompt from {sheet_name} {prompt_cell}: {prompt}")
                        response = client.images.generate(
                            model="dall-e-3",
                            prompt=prompt,
                            size="1024x1024",
                            quality="standard",
                            n=1
                        )
                        image_url = response.data[0].url
                        print("Image URL:", image_url)
                        image_filename = f'{sheet_name}_{image_cell}.png'
                        download_image(image_url, image_filename)
                        insert_image_to_excel(sheet, image_filename, image_cell)
                        break
                    except APIError as e:  # Correctly handle the APIError
                        print(f"Error processing prompt: {e}")
                        attempt += 1
                        if attempt == max_attempts:
                            print("Max attempts reached, moving to next prompt.")

    # Save the workbook with all modifications
    wb.save('output.xlsx')
    print("All modifications saved to 'output.xlsx'.")

if __name__ == "__main__":
    generate_and_insert_images()
